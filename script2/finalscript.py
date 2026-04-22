import argparse
import asyncio
import csv
import json
import os
import random
import re
import time
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

try:
    from gemini_webapi import GeminiClient, set_log_level
    from gemini_webapi.constants import Model
except Exception:
    GeminiClient = None
    set_log_level = None
    Model = None


SYSTEM_PROMPT = """You are an expert Python data processing assistant.
Your task is to write a single, complete Python function named `process_data(rows)`.
The input `rows` is a list of lists (representing rows and columns of an Excel/CSV file).
The first element of the returned list of lists MUST be the clean, final header row.

Rules:
- Write ONLY valid Python code block enclosed in ```python ... ```.
- Use only standard library modules (e.g., `re`, `datetime`). No pandas or third-party libraries.
- The input data might lack clear headers. You must infer column meanings dynamically if needed based on the provided sample data.
- Never output markdown outside the python block. Do not provide explanations.
"""

USER_PROMPT_TEMPLATE = """Sample raw rows (first {sample_count} rows of the file):
{sample_rows}

Standard Processing Guidelines:
1. Column Identification: Systematically identify "EAN", "Name", "Price", and "Stock/Quantity" columns despite varying headers. If headers are missing, infer based on data patterns (e.g., 13 digits = EAN, text = Name, currency/decimals = Price, integers = Quantity).
2. Data Cleaning & Extraction:
   - EAN: Must be strictly a 13-digit string. Left-pad with zeros if shorter. Strip non-numeric characters.
   - Name: Combine brand/manufacturer with the main description if they are in separate columns.
   - Price: Remove currency symbols (e.g., '€', '$'), text, and whitespace. Convert to float. Ensure '.' is used for decimals (e.g., 74.50).
   - Quantity: Remove signs like '+' or text. Convert to integer.
3. Filtering Rules (Exclude rows if any condition is met):
   - Quantity is <= 4 or invalid.
   - Price is < 2.50 or invalid.
    - Total stock value (Price * Quantity) is < 100 EUR.
   - Strict- Row represents incoming stock, delivery dates, or estimated dates , Availability (e.g., "incoming 23.03"). Only keep ready/available stock it can be at any column with any header smartly it has to find and not add if not in stock only in stock has to be added.
{supplier_specific_rules}
4. Calculations: Calculate "Total Price" = (Price * Quantity).
5. Minimum Order Quantity (MOQ): If a minimum order quantity is specified in the list, extract it to a separate column.
6. Final Output Structure: The final valid rows must follow EXACTLY this column order:
   ["EAN", "Name", "Price", "Stock/Quantity", "Total Price", "Supplier"] (plus "Min Qty" as the last column if MOQ info exists).
   - "Supplier" column must be strictly the string "{supplier_name}" for every row.
   - Output integers and floats as actual numeric types in Python so they display as numbers in Excel, except EAN which remains a string.

User specific request:
{instruction}

Generate the full `def process_data(rows: list[list[Any]]) -> list[list[Any]]:` Python code block:
"""

DEFAULT_MODEL_NAME = "gemini-3-pro"

# Cookie values for authentication
Secure_1PSID = "g.a0009AjBI9hSRdlgkzMwaBFwfFO6IRDz7luAOtDwA9ukyvCaGGtJ-QuMcNlLj0XEkU_iAutgRQACgYKAXMSARMSFQHGX2MihMTN1gZhGq2mmGmUreStzxoVAUF8yKoh9C953oCF0ljkS01l1pMx0076"
Secure_1PSIDTS = "sidts-CjEBWhotCSqmH8uPAooyw7E8FPujZzZ-PFbcbNvxz_HdukMG69UKqTSZXTw9uucFu4UJEAA"

if set_log_level is not None:
    set_log_level("INFO")

DEFAULT_MIN_DELAY_SECONDS = 12.0
DEFAULT_MAX_DELAY_SECONDS = 25.0
DEFAULT_MAX_RETRIES = 3
DEFAULT_REQUEST_TIMEOUT_SECONDS = 60
DEFAULT_GENERATION_ATTEMPTS = 3
DEFAULT_CODE_CACHE_DIR = ".code_cache"

AKATRONIK_ALLOWED_BRANDS = [
    "AEG",
    "BEKO",
    "BOSCH",
    "DELONGHI",
    "ELECTROLUX",
    "GORENJE",
    "HISENSE",
    "LG",
    "SAMSUNG",
    "SIEMENS",
]

BIG_APPLIANCE_KEYWORDS = [
    "washing machine",
    "washer",
    "dryer",
    "dishwasher",
    "fridge",
    "refrigerator",
    "freezer",
    "oven",
    "cooker",
    "hob",
    "stove",
]


def get_human_delay(min_delay: float, max_delay: float) -> float:
    if min_delay > max_delay:
        min_delay, max_delay = max_delay, min_delay
    return random.uniform(min_delay, max_delay)


def get_tokens(args: argparse.Namespace) -> tuple[str, str]:
    secure_1psid = args.secure_1psid or os.getenv("SECURE_1PSID", Secure_1PSID)
    secure_1psidts = args.secure_1psidts or os.getenv("SECURE_1PSIDTS", Secure_1PSIDTS)
    if not secure_1psid or not secure_1psidts:
        raise ValueError("Missing PSID tokens.")
    return secure_1psid, secure_1psidts


def resolve_model(model_name: str):
    if Model is None:
        raise RuntimeError("gemini_webapi is not installed.")
    if hasattr(Model, model_name):
        return getattr(Model, model_name)
    return model_name


def _is_stalled_or_timeout_error(exc: Exception) -> bool:
    if isinstance(exc, asyncio.TimeoutError):
        return True
    message = str(exc).lower()
    stall_markers = [
        "response stalled",
        "no progress",
        "timed out",
        "timeout",
        "queueing=false",
        "queueing = false",
    ]
    return any(marker in message for marker in stall_markers)


async def init_gemini_client(args: argparse.Namespace):
    if GeminiClient is None:
        raise RuntimeError("gemini_webapi module not found. Install it.")

    secure_1psid, secure_1psidts = get_tokens(args)
    client = GeminiClient(secure_1psid, secure_1psidts, proxy=args.proxy)
    await client.init(timeout=args.gemini_timeout, auto_close=False, close_delay=300, auto_refresh=True)
    return client


async def close_gemini_client(client: Any) -> None:
    if client is None:
        return
    close_method = getattr(client, "close", None)
    if close_method is None:
        return
    result = close_method()
    if asyncio.iscoroutine(result):
        await result


async def generate_code_with_gemini(
    system_prompt: str,
    user_prompt: str,
    args: argparse.Namespace,
    client: Any,
) -> str:
    if client is None:
        raise RuntimeError("Gemini client is not initialized.")

    final_prompt = f"{system_prompt}\n\n{user_prompt}"
    last_error = None

    for attempt in range(1, args.max_retries + 1):
        chat = None
        try:
            chat = client.start_chat(model=resolve_model(args.gemini_model))
            if attempt > 1:
                await asyncio.sleep(get_human_delay(args.min_delay, args.max_delay))
            
            response = await asyncio.wait_for(
                chat.send_message(final_prompt),
                timeout=args.request_timeout,
            )
            text = (response.text or "").strip()
            
            if not text:
                raise ValueError("Gemini returned empty response text")
                
            code_match = re.search(r"```(?:python)?\s*(.*?)```", text, re.DOTALL | re.IGNORECASE)
            if code_match:
                return code_match.group(1).strip()
            return text 
            
        except Exception as exc:
            last_error = exc
            if attempt < args.max_retries:
                if _is_stalled_or_timeout_error(exc):
                    print(
                        f"Gemini stalled/timed out on attempt {attempt}/{args.max_retries}. "
                        "Retrying with a fresh chat immediately..."
                    )
                else:
                    await asyncio.sleep(get_human_delay(args.min_delay, args.max_delay))
        finally:
            if chat is not None:
                close_method = getattr(chat, "close", None)
                if close_method is not None:
                    result = close_method()
                    if asyncio.iscoroutine(result):
                        await result

    raise RuntimeError(f"Gemini request failed: {last_error}")


def _supplier_cache_key(supplier_name: str) -> str:
    key = re.sub(r"[^A-Za-z0-9._-]+", "_", supplier_name).strip("_")
    return key or "supplier"


def _read_cached_code(cache_dir: Path, supplier_name: str) -> str:
    cache_path = cache_dir / f"{_supplier_cache_key(supplier_name)}.py"
    if cache_path.exists() and cache_path.is_file():
        return cache_path.read_text(encoding="utf-8")
    return ""


def _write_cached_code(cache_dir: Path, supplier_name: str, python_code: str) -> None:
    if not python_code.strip():
        return
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_path = cache_dir / f"{_supplier_cache_key(supplier_name)}.py"
    cache_path.write_text(python_code, encoding="utf-8")


def _generate_code_with_fallback(
    system_prompt: str,
    user_prompt: str,
    args: argparse.Namespace,
    supplier_name: str,
    gemini_loop: asyncio.AbstractEventLoop | None,
    gemini_client: Any,
) -> str:
    cache_dir = Path(args.code_cache_dir)
    last_error: Exception | None = None

    if gemini_loop is None or gemini_client is None:
        raise RuntimeError("Shared Gemini client/event loop is not available.")

    for attempt in range(1, args.generation_attempts + 1):
        try:
            print(f"Gemini generation attempt {attempt}/{args.generation_attempts}...")
            python_code = gemini_loop.run_until_complete(
                generate_code_with_gemini(system_prompt, user_prompt, args, gemini_client)
            )
            if not python_code.strip():
                raise ValueError("Gemini returned empty python code")

            _write_cached_code(cache_dir, supplier_name, python_code)
            return python_code
        except Exception as exc:
            last_error = exc
            print(f"Attempt {attempt} failed for supplier '{supplier_name}': {exc}")
            if attempt < args.generation_attempts:
                sleep_seconds = get_human_delay(args.min_delay, args.max_delay)
                print(f"Cooling down for {sleep_seconds:.1f}s before retry...")
                time.sleep(sleep_seconds)

    if not args.no_cache_fallback:
        cached_code = _read_cached_code(cache_dir, supplier_name)
        if cached_code.strip():
            print(f"Using cached code for supplier '{supplier_name}' after Gemini failures.")
            return cached_code

    raise RuntimeError(f"Gemini API failed and no cache fallback available: {last_error}")


def read_excel_raw(path: Path, sheet: str | int | None) -> list[list[Any]]:
    if path.suffix.lower() == ".xls":
        import xlrd
        wb = xlrd.open_workbook(path)
        if sheet is None:
            ws = wb.sheet_by_index(0)
        elif isinstance(sheet, int):
            ws = wb.sheet_by_index(sheet)
        else:
            ws = wb.sheet_by_name(sheet)
        return [ws.row_values(i) for i in range(ws.nrows)]

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]] if sheet is None else (wb[wb.sheetnames[sheet]] if isinstance(sheet, int) else wb[sheet])
    return [list(row) for row in ws.iter_rows(values_only=True)]

def read_csv_raw(path: Path) -> list[list[Any]]:
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        return [row for row in csv.reader(f)]

def read_data_raw(path: Path, sheet: str | int | None) -> list[list[Any]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}:
        return read_excel_raw(path, sheet)
    if suffix == ".csv":
        return read_csv_raw(path)
    raise ValueError("Unsupported input format.")


def write_excel_raw(path: Path, data: list[list[Any]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "clean_output"
    for row in data:
        ws.append(row)

    if data and isinstance(data[0], list):
        header = [str(h).strip().lower().replace(" ", "") for h in data[0]]
        for col_index, col_name in enumerate(header, start=1):
            if col_name in {"price", "totalprice"}:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_index).number_format = "0.00"

    wb.save(path)


def write_csv_raw(path: Path, data: list[list[Any]], delimiter: str = ","):
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=delimiter)
        writer.writerows(data)


def export_data_raw(data: list[list[Any]], out_path: Path, out_format: str):
    if not data:
        return
    fmt = out_format.lower()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if fmt == "excel":
        write_excel_raw(out_path, data)
    elif fmt == "csv":
        write_csv_raw(out_path, data, delimiter=",")
    elif fmt in {"csv_pipe", "pipe_csv"}:
        write_csv_raw(out_path, data, delimiter="|")
    else:
        raise ValueError("Invalid output format")


def _normalize_brand_text(text: Any) -> str:
    if text is None:
        return ""
    value = str(text).strip().upper()
    value = value.replace("'", "")
    return re.sub(r"[^A-Z0-9 ]+", " ", value)


def _contains_allowed_akatronik_brand(name: Any) -> bool:
    normalized = f" {_normalize_brand_text(name)} "
    return any(f" {brand} " in normalized for brand in AKATRONIK_ALLOWED_BRANDS)


def _contains_big_appliance_keyword(name: Any) -> bool:
    normalized = str(name or "").strip().lower()
    if not normalized:
        return False
    return any(keyword in normalized for keyword in BIG_APPLIANCE_KEYWORDS)


def _enforce_price_rounding(processed_data: list[list[Any]]) -> list[list[Any]]:
    if not processed_data:
        return processed_data

    header = processed_data[0]
    if not isinstance(header, list):
        return processed_data

    normalized_header = [str(h).strip().lower().replace(" ", "") for h in header]
    price_cols = [
        idx for idx, col_name in enumerate(normalized_header) if col_name in {"price", "totalprice"}
    ]
    if not price_cols:
        return processed_data

    rounded_rows = [header]
    for row in processed_data[1:]:
        if not isinstance(row, list):
            continue
        row_copy = list(row)
        for idx in price_cols:
            if idx >= len(row_copy):
                continue
            try:
                row_copy[idx] = round(float(row_copy[idx]), 2)
            except (TypeError, ValueError):
                continue
        rounded_rows.append(row_copy)
    return rounded_rows


def _apply_supplier_guardrail_filters(processed_data: list[list[Any]]) -> list[list[Any]]:
    if not processed_data:
        return processed_data

    header = processed_data[0]
    if not isinstance(header, list):
        return processed_data

    normalized_header = [str(h).strip().lower().replace(" ", "") for h in header]
    try:
        supplier_idx = normalized_header.index("supplier")
        name_idx = normalized_header.index("name")
    except ValueError:
        return processed_data

    filtered_rows = [header]
    for row in processed_data[1:]:
        if not isinstance(row, list):
            continue
        if max(supplier_idx, name_idx) >= len(row):
            continue

        supplier = str(row[supplier_idx] or "").strip().lower()
        name = row[name_idx]

        if supplier in {"akatronik", "akatronic"} and not _contains_allowed_akatronik_brand(name):
            continue

        if supplier in {"akatronik", "akatronic", "duna"} and _contains_big_appliance_keyword(name):
            continue

        filtered_rows.append(row)

    return filtered_rows


def _dedupe_by_ean_lowest_price(processed_data: list[list[Any]]) -> list[list[Any]]:
    """Keep one row per EAN, selecting lowest Price and highest stock for ties."""
    if not processed_data or len(processed_data) <= 2:
        return processed_data

    header = processed_data[0]
    if not isinstance(header, list):
        return processed_data

    header_norm = [str(h).strip().lower().replace(" ", "") for h in header]

    try:
        ean_idx = header_norm.index("ean")
        price_idx = header_norm.index("price")
    except ValueError:
        return processed_data

    qty_idx = None
    if "stock/quantity" in header_norm:
        qty_idx = header_norm.index("stock/quantity")

    def _to_float(value: Any) -> float:
        try:
            return float(value)
        except (TypeError, ValueError):
            return float("inf")

    def _to_qty(value: Any) -> float:
        try:
            return float(value)
        except (TypeError, ValueError):
            return float("-inf")

    best_by_ean: dict[str, list[Any]] = {}
    passthrough_rows: list[list[Any]] = []

    for row in processed_data[1:]:
        if not isinstance(row, list):
            continue
        if max(ean_idx, price_idx) >= len(row):
            continue

        ean = str(row[ean_idx] or "").strip()
        if not ean:
            passthrough_rows.append(row)
            continue

        price = _to_float(row[price_idx])
        qty = _to_qty(row[qty_idx]) if qty_idx is not None and qty_idx < len(row) else float("-inf")

        current = best_by_ean.get(ean)
        if current is None:
            best_by_ean[ean] = row
            continue

        current_price = _to_float(current[price_idx]) if price_idx < len(current) else float("inf")
        current_qty = _to_qty(current[qty_idx]) if qty_idx is not None and qty_idx < len(current) else float("-inf")

        if (price < current_price) or (price == current_price and qty > current_qty):
            best_by_ean[ean] = row

    return [header, *best_by_ean.values(), *passthrough_rows]


def format_prompt(sample_rows: list[list[Any]], instruction: str, supplier_name: str) -> str:
    sample_str = ""
    for idx, row in enumerate(sample_rows):
        sample_str += f"Row {idx}: {row}\n"
        
    supplier_rules = ""
    supplier_key = supplier_name.lower().strip()
    if supplier_key in {"akatronik", "akatronic"}:
        supplier_rules = "   - STRICT AKATRONIK FILTER: You MUST ONLY KEEP items where the Name contains one of these exact brands (ignoring case): AEG, BEKO, BOSCH, De'Longhi, ELECTROLUX, Gorenje, Hisense, LG, SAMSUNG, Siemens. Otherwise, filter out the entire row."
    elif supplier_key == "duna":
        supplier_rules = "   - STRICT DUNA FILTER: Exclude obvious large appliances from Name text (washing machine, dryer, dishwasher, refrigerator/freezer, oven/hob/cooker/stove)."
        
    return USER_PROMPT_TEMPLATE.format(
        sample_count=len(sample_rows),
        sample_rows=sample_str,
        instruction=instruction,
        supplier_name=supplier_name,
        supplier_specific_rules=supplier_rules
    )


def extract_code_from_pasted_input() -> str:
    lines = []
    print("\nPaste the Gemini Python code block below. End with an empty line (two enters):\n")
    while True:
        line = input()
        if line.strip() == "":
            break
        lines.append(line)
    text = "\n".join(lines)
    code_match = re.search(r"```(?:python)?\s*(.*?)```", text, re.DOTALL | re.IGNORECASE)
    if code_match:
        return code_match.group(1).strip()
    return text.strip()


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Smart Excel/CSV processor powered by LLM code generation (Folder batch mode).")
    p.add_argument("--input-folder", default="input", help="Folder containing .xlsx/.xlsm/.csv files")
    p.add_argument("--sheet", help="Excel sheet name or index")
    p.add_argument("--output", default="output.xlsx", help="Output file path")
    p.add_argument("--output-format", default="excel", choices=["excel", "csv", "csv_pipe"])

    p.add_argument("--instruction", help="Natural language extraction/filter request", default="")
    p.add_argument("--code-file", help="Path to already generated python code file to execute")
    p.add_argument("--save-code-file", help="Optional prefix to save generated python code locally")
    
    p.add_argument("--print-prompts", action="store_true", help="Print SYSTEM + USER prompt and exit")
    p.add_argument("--interactive", action="store_true", help="Interactive mode for script paste")
    
    p.add_argument("--secure-1psid", help="Gemini web token Secure_1PSID")
    p.add_argument("--secure-1psidts", help="Gemini web token Secure_1PSIDTS")
    p.add_argument("--gemini-model", default=DEFAULT_MODEL_NAME)
    p.add_argument("--gemini-timeout", type=int, default=350)
    p.add_argument("--request-timeout", type=int, default=DEFAULT_REQUEST_TIMEOUT_SECONDS,
                   help="Max seconds to wait for a single Gemini send_message call")
    p.add_argument("--proxy", default=None)
    p.add_argument("--min-delay", type=float, default=DEFAULT_MIN_DELAY_SECONDS)
    p.add_argument("--max-delay", type=float, default=DEFAULT_MAX_DELAY_SECONDS)
    p.add_argument("--max-retries", type=int, default=DEFAULT_MAX_RETRIES)
    p.add_argument("--generation-attempts", type=int, default=DEFAULT_GENERATION_ATTEMPTS,
                   help="How many full generation attempts to run per supplier")
    p.add_argument("--code-cache-dir", default=DEFAULT_CODE_CACHE_DIR,
                   help="Directory to store/reuse generated supplier parser code")
    p.add_argument("--no-cache-fallback", action="store_true",
                   help="Disable fallback to cached supplier code when Gemini fails")
    
    return p


def run_pipeline(args: argparse.Namespace):
    input_folder = Path(args.input_folder)
    
    if not input_folder.exists() or not input_folder.is_dir():
        print(f"Creating folder '{input_folder}' because it wasn't found.")
        input_folder.mkdir(parents=True, exist_ok=True)
        print(f"Please put your Excel/CSV files in the '{input_folder}' directory and rerun the script.")
        return

    all_files = [f for f in input_folder.iterdir() if f.suffix.lower() in {'.xlsx', '.xls', '.xlsm', '.csv'}]
    
    if not all_files:
        print(f"No valid data files found in folder: {input_folder}")
        return

    if not args.instruction and not args.code_file and not args.interactive and not args.print_prompts:
        args.instruction = input("\nEnter any specific instructions (or press Enter to just apply standard rules): ").strip()

    master_rows = []
    headers = None
    
    status_report = []
    gemini_loop: asyncio.AbstractEventLoop | None = None
    gemini_client = None
    use_gemini_generation = not args.code_file and not args.interactive and not args.print_prompts

    if use_gemini_generation:
        try:
            gemini_loop = asyncio.new_event_loop()
            asyncio.set_event_loop(gemini_loop)
            gemini_client = gemini_loop.run_until_complete(init_gemini_client(args))
            print("Gemini client initialized once for this full batch.")
        except Exception as e:
            print(f"Failed to initialize Gemini client: {e}")
            if gemini_loop is not None:
                gemini_loop.close()
                asyncio.set_event_loop(None)
            return

    print(f"\nFound {len(all_files)} file(s) in '{input_folder}'. Starting batch processing...\n")

    try:
        for file_idx, file_path in enumerate(all_files):
            supplier_name = file_path.stem
            print(f"{'='*60}")
            print(f"Processing: {file_path.name} (Supplier: {supplier_name})")
            print(f"{'='*60}")

            sheet = int(args.sheet) if args.sheet and args.sheet.isdigit() else args.sheet

            try:
                raw_data = read_data_raw(file_path, sheet)
            except Exception as e:
                print(f"Failed to read {file_path.name}: {e}")
                status_report.append({"file": file_path.name, "status": "Failed to Read", "rows": 0})
                continue

            if not raw_data:
                print(f"Warning: {file_path.name} is empty.")
                status_report.append({"file": file_path.name, "status": "Empty File", "rows": 0})
                continue

            sample_rows = raw_data[:15]
            prompt = format_prompt(sample_rows, args.instruction, supplier_name)

            if args.print_prompts:
                print(f"\n===== PROMPTS FOR {file_path.name} =====")
                print("System Prompt Length:", len(SYSTEM_PROMPT))
                print("User Prompt Template:\n")
                print(prompt)
                status_report.append({"file": file_path.name, "status": "Printed Prompt", "rows": 0})
                continue

            python_code = ""

            if args.code_file:
                python_code = Path(args.code_file).read_text(encoding="utf-8")
            elif args.interactive:
                python_code = extract_code_from_pasted_input()
            else:
                print("Requesting code generation from Gemini for this specific file format...")
                try:
                    python_code = _generate_code_with_fallback(
                        SYSTEM_PROMPT,
                        prompt,
                        args,
                        supplier_name,
                        gemini_loop,
                        gemini_client,
                    )

                    if args.save_code_file:
                        save_path = Path(f"{args.save_code_file}_{supplier_name}.py")
                        save_path.parent.mkdir(parents=True, exist_ok=True)
                        save_path.write_text(python_code, encoding="utf-8")

                except Exception as e:
                    print(f"Gemini API Error for {file_path.name}: {e}")
                    status_report.append({"file": file_path.name, "status": "LLM API Error", "rows": 0})
                    continue

            if not python_code.strip():
                print(f"No code generated for {file_path.name}")
                status_report.append({"file": file_path.name, "status": "Failed - No Code", "rows": 0})
                continue

            print("Executing dynamically generated Python filter...")
            exec_globals = {}
            try:
                exec(python_code, exec_globals)
                process_func = exec_globals.get("process_data")

                if not process_func:
                    raise KeyError("The generated code did not define a 'process_data' function.")

                processed_data = process_func(raw_data)

                # Safety net: enforce minimum stock-value rule after generated processing.
                if processed_data and len(processed_data) > 1:
                    header = processed_data[0]
                    header_norm = [str(h).strip().lower().replace(" ", "") for h in header]
                    try:
                        price_idx = header_norm.index("price")
                        qty_idx = header_norm.index("stock/quantity")
                    except ValueError:
                        price_idx = None
                        qty_idx = None

                    if price_idx is not None and qty_idx is not None:
                        filtered_rows = [header]
                        for row in processed_data[1:]:
                            if not isinstance(row, list):
                                continue
                            if max(price_idx, qty_idx) >= len(row):
                                continue
                            try:
                                price = float(row[price_idx])
                                qty = float(row[qty_idx])
                            except (TypeError, ValueError):
                                continue
                            if (price * qty) >= 100:
                                filtered_rows.append(row)
                        processed_data = filtered_rows

                processed_data = _apply_supplier_guardrail_filters(processed_data)
                processed_data = _enforce_price_rounding(processed_data)

                if not processed_data or len(processed_data) < 2:
                    print(f"Skipped {file_path.name}: Output data was empty after filtering.")
                    status_report.append({"file": file_path.name, "status": "Success (No matches)", "rows": 0})
                    continue

                current_headers = processed_data[0]
                if not headers:
                    headers = current_headers
                    master_rows.append(headers)

                data_rows_only = processed_data[1:]
                master_rows.extend(data_rows_only)
                print(f"✓ Extracted {len(data_rows_only)} valid products.")
                status_report.append({"file": file_path.name, "status": "Success", "rows": len(data_rows_only)})

            except Exception as e:
                print("=== GENERATED SCRIPT CRASHED ===")
                print(python_code)
                print("================================")
                print(f"Execution Error on {file_path.name}: {e}")
                status_report.append({"file": file_path.name, "status": "Script Execution Error", "rows": 0})

            if file_idx < len(all_files) - 1:
                sleep_seconds = get_human_delay(args.min_delay, args.max_delay)
                print(f"Waiting {sleep_seconds:.1f}s before next supplier to reduce stream throttling...")
                time.sleep(sleep_seconds)

        if not args.print_prompts:
            print("\n" + "="*80)
            print(" FINAL PIPELINE SUMMARY ")
            print("="*80)
            print(f"{'File Name':<30} | {'Status':<25} | {'Extracted Rows':<15}")
            print("-" * 80)

            total_rows = 0
            for item in status_report:
                print(f"{item['file'][:28]:<30} | {item['status'][:23]:<25} | {item['rows']:<15}")
                if item['rows']:
                    total_rows += int(item['rows'])

            print("-" * 80)
            print(f"Total Combined Clean Rows: {total_rows}")

            if len(master_rows) > 1:
                master_rows = _dedupe_by_ean_lowest_price(master_rows)
                export_data_raw(master_rows, Path(args.output), args.output_format)
                print(f"\nSUCCESS: All data consolidated and saved to -> {args.output}")
            else:
                print("\nWARNING: No valid rows were filtered from any files (or no header obtained). No output saved.")
    finally:
        if gemini_loop is not None:
            try:
                gemini_loop.run_until_complete(close_gemini_client(gemini_client))
            except Exception as close_error:
                print(f"Warning: failed to close Gemini client cleanly: {close_error}")
            gemini_loop.close()
            asyncio.set_event_loop(None)

def main():
    parser = build_parser()
    args = parser.parse_args()
    run_pipeline(args)


if __name__ == "__main__":
    main()
