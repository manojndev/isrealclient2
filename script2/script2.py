from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_FILE = BASE_DIR / "output.xlsx"
DATABASE_FILE = BASE_DIR / "database.xlsx"
FINAL_OUTPUT_FILE = BASE_DIR / "finaloutput.xlsx"
BOL_OUTPUT_FILE = BASE_DIR / "bol.xlsx"
DIGITEC_OUTPUT_FILE = BASE_DIR / "digitec.xlsx"

AKATRONIK_DISALLOWED_BRANDS = [
	"AEG",
	"BEKO",
	"BOSCH",
	"DELONGHI",
	"ELECTROLUX",
	"GORENJE",
	"HISENSE",
	"LG",
	"SAMSUNG",
	"SIEMENS",
]

BIG_APPLIANCE_KEYWORDS = [
	"washing machine",
	"washer",
	"dryer",
	"dishwasher",
	"fridge",
	"refrigerator",
	"freezer",
	"oven",
	"cooker",
	"hob",
	"stove",
]

APPLE_IOS_DEVICE_REGEX = re.compile(r"\b(iphone|ipad)\b", re.IGNORECASE)


STATIC_DIGITEC_ROWS = [
	{
		"Gtin": "8720389022166",
		"ProviderKey": "P28398",
		"ManufacturerKey": "",
		"BrandName": "PHILIPS",
		"ProductTitle": "PHILIPS HD9255/30",
		"PurchasePriceExclVat": 79.00,
		"Currency": "EURO",
		"QuantityOnStock": 10,
		"MinimumOrderQuantity": 10,
		"Order Quantity Steps": 10,
	},
	{
		"Gtin": "5999024868329",
		"ProviderKey": "P17424",
		"ManufacturerKey": "",
		"BrandName": "Dometic",
		"ProductTitle": "Dometic ACX3 40 G CombiCool tragbare Absorber-Kuhlbox",
		"PurchasePriceExclVat": 250.00,
		"Currency": "EURO",
		"QuantityOnStock": 9,
		"MinimumOrderQuantity": 4,
		"Order Quantity Steps": 4,
	},
	{
		"Gtin": "6934177746512",
		"ProviderKey": "P7571",
		"ManufacturerKey": "",
		"BrandName": "Xiaomi",
		"ProductTitle": "Xiaomi Auricolari Wireless Redmi Buds 3 Pro Glacier Gray",
		"PurchasePriceExclVat": 10.00,
		"Currency": "EURO",
		"QuantityOnStock": 50,
		"MinimumOrderQuantity": 30,
		"Order Quantity Steps": 30,
	},
	{
		"Gtin": "4001627025465",
		"ProviderKey": "P2186",
		"ManufacturerKey": "",
		"BrandName": "GRAEF",
		"ProductTitle": "GRAEF TB 501 STAND BLENDER",
		"PurchasePriceExclVat": 31.00,
		"Currency": "EURO",
		"QuantityOnStock": 10,
		"MinimumOrderQuantity": 10,
		"Order Quantity Steps": 10,
	},
	{
		"Gtin": "6925281982095",
		"ProviderKey": "P2951",
		"ManufacturerKey": "",
		"BrandName": "JBL",
		"ProductTitle": "JBL Charge 5 Blue",
		"PurchasePriceExclVat": 63.00,
		"Currency": "EURO",
		"QuantityOnStock": 1,
		"MinimumOrderQuantity": 1,
		"Order Quantity Steps": 1,
	},
	{
		"Gtin": "4054278497143",
		"ProviderKey": "P2463",
		"ManufacturerKey": "",
		"BrandName": "Karcher",
		"ProductTitle": "Karcher K7 Comact Home 1.447-053.0",
		"PurchasePriceExclVat": 291.00,
		"Currency": "EURO",
		"QuantityOnStock": 10,
		"MinimumOrderQuantity": 3,
		"Order Quantity Steps": 3,
	},
	{
		"Gtin": "0088381779210",
		"ProviderKey": "P32862",
		"ManufacturerKey": "",
		"BrandName": "Makita",
		"ProductTitle": "DTD173Z Akku-Schlagschrauber LXT",
		"PurchasePriceExclVat": 128.00,
		"Currency": "EURO",
		"QuantityOnStock": 10,
		"MinimumOrderQuantity": 6,
		"Order Quantity Steps": 6,
	},
]


def normalize_gtin(value: object) -> str:
	"""Normalize GTIN/EAN values to a 13-digit string for reliable matching."""
	if pd.isna(value):
		return ""

	text = str(value).strip()
	if not text:
		return ""

	if "." in text:
		text = text.split(".", 1)[0]

	digits = "".join(ch for ch in text if ch.isdigit())
	if not digits:
		return ""

	if len(digits) > 13:
		return digits[-13:]
	return digits.zfill(13)


def pick_column(df: pd.DataFrame, candidates: list[str]) -> str:
	normalized = {
		str(col).strip().lower().replace(" ", ""): col for col in df.columns
	}
	for candidate in candidates:
		key = candidate.strip().lower().replace(" ", "")
		if key in normalized:
			return normalized[key]
	raise KeyError(f"None of these columns were found: {candidates}")


def split_code(code: object) -> tuple[str, int, int] | None:
	"""Split a product code like P41226 into prefix, numeric part, and numeric width."""
	if pd.isna(code):
		return None

	text = str(code).strip()
	if not text:
		return None

	match = re.match(r"^([A-Za-z]+)(\d+)$", text)
	if not match:
		return None

	prefix = match.group(1)
	number_text = match.group(2)
	return prefix, int(number_text), len(number_text)


def get_last_product_code(db_df: pd.DataFrame, code_col: str) -> tuple[str, int, int]:
	"""Find the last valid product code in database order."""
	for value in reversed(db_df[code_col].tolist()):
		parts = split_code(value)
		if parts is not None:
			return parts

	return "P", 0, 1


def derive_brand(text: object) -> str:
	"""Extract a simple brand from the first token of a product title."""
	if pd.isna(text):
		return ""

	raw = str(text).strip()
	if not raw:
		return ""

	first = raw.split()[0]
	clean = "".join(ch for ch in first if ch.isalnum())
	return clean or first


def calculate_hifi_price(selling_price: object) -> object:
	"""Apply the requested tiered markup formula based on selling price."""
	if pd.isna(selling_price):
		return ""

	try:
		price = float(selling_price)
	except (TypeError, ValueError):
		return ""

	if price <= 4:
		multiplier = 1.70
	elif price <= 6:
		multiplier = 1.65
	elif price <= 10:
		multiplier = 1.50
	elif price <= 15:
		multiplier = 1.40
	elif price <= 20:
		multiplier = 1.30
	elif price <= 30:
		multiplier = 1.25
	elif price <= 50:
		multiplier = 1.19
	elif price <= 90:
		multiplier = 1.15
	elif price <= 150:
		multiplier = 1.14
	elif price <= 190:
		multiplier = 1.12
	elif price <= 270:
		multiplier = 1.11
	elif price <= 500:
		multiplier = 1.10
	elif price <= 800:
		multiplier = 1.105
	elif price <= 5000:
		multiplier = 1.10
	else:
		return ""

	return round(price * multiplier, 2)


def is_apple_iphone_or_ipad(brand: object, title: object) -> bool:
	"""Match Apple iPhone/iPad products to apply dedicated markup."""
	brand_text = "" if pd.isna(brand) else str(brand).strip().lower()
	title_text = "" if pd.isna(title) else str(title).strip().lower()
	combined = f"{brand_text} {title_text}".strip()
	if not combined:
		return False
	if "apple" not in combined:
		return False
	return bool(APPLE_IOS_DEVICE_REGEX.search(combined))


def calculate_hifi_price_with_overrides(selling_price: object, brand: object, title: object) -> object:
	"""Apply product-specific overrides before default pricing tiers."""
	base_price = to_number(selling_price)
	if base_price is None or base_price <= 0:
		return ""

	if is_apple_iphone_or_ipad(brand, title):
		return round(base_price * 1.045, 2)

	return calculate_hifi_price(selling_price)


def dedupe_by_ean_lowest_price(
	df: pd.DataFrame,
	ean_col: str,
	price_col: str,
	qty_col: str,
) -> pd.DataFrame:
	"""Keep one row per EAN, selecting the lowest price (then highest stock on ties)."""
	if df.empty:
		return df

	work = df.copy()
	work["_ean_norm"] = work[ean_col].apply(normalize_gtin)
	work["_price_sort"] = pd.to_numeric(work[price_col], errors="coerce")
	work["_qty_sort"] = pd.to_numeric(work[qty_col], errors="coerce")

	# Keep rows with valid EAN and price first to maximize deterministic selection quality.
	with_valid_ean = work[work["_ean_norm"] != ""].copy()
	without_valid_ean = work[work["_ean_norm"] == ""].copy()

	if not with_valid_ean.empty:
		with_valid_ean = with_valid_ean.sort_values(
			by=["_ean_norm", "_price_sort", "_qty_sort"],
			ascending=[True, True, False],
			na_position="last",
		)
		with_valid_ean = with_valid_ean.drop_duplicates(subset=["_ean_norm"], keep="first")

	combined = pd.concat([with_valid_ean, without_valid_ean], ignore_index=True)
	return combined.drop(columns=["_ean_norm", "_price_sort", "_qty_sort"], errors="ignore")


def to_number(value: object) -> float | None:
	"""Convert values to float when possible, else return None."""
	if pd.isna(value):
		return None
	try:
		return float(value)
	except (TypeError, ValueError):
		return None


def calculate_min_quantity(sum_available: object, hifi_price: object) -> object:
	"""Cap Min quantity to stock<=25 and keep Min quantity*Hifi price <= 1000."""
	qty_num = to_number(sum_available)
	price_num = to_number(hifi_price)

	if qty_num is None or qty_num <= 0:
		return 0

	base_qty = int(min(qty_num, 25))
	if price_num is None or price_num <= 0:
		return base_qty

	max_qty_under_1000 = int(1000 // price_num)
	if max_qty_under_1000 >= 1:
		return min(base_qty, max_qty_under_1000)

	# If one unit is already above 1000, keep at least 1 instead of dropping to 0.
	return 1


def normalize_brand_text(text: object) -> str:
	if pd.isna(text):
		return ""
	value = str(text).strip().upper().replace("'", "")
	return re.sub(r"[^A-Z0-9 ]+", " ", value)


def contains_disallowed_akatronik_brand(name: object) -> bool:
	normalized = f" {normalize_brand_text(name)} "
	return any(f" {brand} " in normalized for brand in AKATRONIK_DISALLOWED_BRANDS)


def contains_big_appliance_keyword(name: object) -> bool:
	if pd.isna(name):
		return False
	text = str(name).strip().lower()
	if not text:
		return False
	return any(keyword in text for keyword in BIG_APPLIANCE_KEYWORDS)


def apply_supplier_filters(df: pd.DataFrame, supplier_col: str, name_col: str) -> pd.DataFrame:
	def keep_row(row: pd.Series) -> bool:
		supplier = str(row.get(supplier_col, "") or "").strip().lower()
		name = row.get(name_col, "")

		if supplier in {"akatronik", "akatronic"} and contains_disallowed_akatronik_brand(name):
			return False

		if supplier in {"akatronik", "akatronic", "duna"} and contains_big_appliance_keyword(name):
			return False

		return True

	return df[df.apply(keep_row, axis=1)].copy()


def apply_text_format_to_barcode(path: Path) -> None:
	wb = load_workbook(path)
	for sheet_name in ["finaloutput", "new products"]:
		if sheet_name not in wb.sheetnames:
			continue
		ws = wb[sheet_name]
		for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
			row[0].number_format = "@"
			if len(row) > 1:
				row[1].number_format = "@"

		if sheet_name == "finaloutput":
			for col in (5, 7, 10):
				for row_idx in range(2, ws.max_row + 1):
					ws.cell(row=row_idx, column=col).number_format = "0.00"

	if "Aanbod" in wb.sheetnames:
		ws = wb["Aanbod"]
		# Internal reference (col B) and Product reference (col C) should stay text.
		for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
			for cell in row:
				cell.number_format = "@"
		for row_idx in range(2, ws.max_row + 1):
			ws.cell(row=row_idx, column=8).number_format = "0.00"

	if "digitec" in wb.sheetnames:
		ws = wb["digitec"]
		# GTIN (col A) and ProviderKey (col B) should stay text.
		for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
			for cell in row:
				cell.number_format = "@"
		for row_idx in range(2, ws.max_row + 1):
			ws.cell(row=row_idx, column=6).number_format = "0.00"
	wb.save(path)


def build_bol_dataframe(final_df: pd.DataFrame) -> pd.DataFrame:
	"""Build BOL export with requested schema and field mapping."""
	bol_columns = [
		"Catalog",
		"Internal reference",
		"Product reference",
		"Name",
		"Product classification",
		"List price",
		"Published price to dealer",
		"Net purchase price",
		"Currency",
		"VAT",
		"Max delivery time",
		"Order quantity",
		"In stock",
		"Quantity stock",
		"Release date",
		"On backorder",
		"Quantity on backorder",
		"Replenishment date",
		"Packing unit",
		"NPP carton",
		"Quantity carton",
		"NPP layer",
		"Quantity layer",
		"NPP pallet",
		"Quantity pallet",
		"NPP truck",
		"Quantity truck",
		"Fulfilment supplier",
		"Fulfilment ID",
	]

	bol_df = pd.DataFrame(index=final_df.index, columns=bol_columns)
	bol_df = bol_df.fillna("")

	bol_df["Internal reference"] = final_df["Hifi code"]
	bol_df["Product reference"] = final_df["barcode"]
	bol_df["Name"] = final_df["ProductTitle"]
	bol_df["Net purchase price"] = final_df["Hifi price"]
	bol_df["Currency"] = "EUR"
	bol_df["VAT"] = 21
	bol_df["Max delivery time"] = 15
	bol_df["Order quantity"] = final_df["Min quantity"]
	bol_df["In stock"] = "Y"
	bol_df["Quantity stock"] = final_df["sum Available"]

	return bol_df


def build_digitec_dataframe(final_df: pd.DataFrame) -> pd.DataFrame:
	"""Build Digitec export with requested schema and field mapping."""
	digitec_columns = [
		"Gtin",
		"ProviderKey",
		"ManufacturerKey",
		"BrandName",
		"ProductTitle",
		"PurchasePriceExclVat",
		"Currency",
		"QuantityOnStock",
		"MinimumOrderQuantity",
		"Order Quantity Steps",
	]

	digitec_df = pd.DataFrame(index=final_df.index, columns=digitec_columns)
	digitec_df = digitec_df.fillna("")

	digitec_df["Gtin"] = final_df["barcode"]
	digitec_df["ProviderKey"] = final_df["Hifi code"]
	digitec_df["ManufacturerKey"] = ""
	digitec_df["BrandName"] = final_df["Name"]
	digitec_df["ProductTitle"] = final_df["ProductTitle"]
	digitec_df["PurchasePriceExclVat"] = final_df["Hifi price"]
	digitec_df["Currency"] = "EURO"
	digitec_df["QuantityOnStock"] = final_df["sum Available"]
	digitec_df["MinimumOrderQuantity"] = final_df["Min quantity"]
	digitec_df["Order Quantity Steps"] = final_df["Min quantity"]

	# Always append required static Digitec rows at the bottom.
	static_df = pd.DataFrame(STATIC_DIGITEC_ROWS, columns=digitec_columns)
	digitec_df = pd.concat([digitec_df, static_df], ignore_index=True)

	return digitec_df


def main() -> None:
	output_df = pd.read_excel(OUTPUT_FILE, sheet_name="clean_output")
	db_df = pd.read_excel(DATABASE_FILE, sheet_name="DataBase")

	output_ean_col = pick_column(output_df, ["EAN", "barcode", "gtin"])
	output_name_col = pick_column(output_df, ["Name", "ProductTitle", "title"])
	output_price_col = pick_column(output_df, ["Price", "selling Price", "Selling price"])
	output_qty_col = pick_column(output_df, ["Stock/Quantity", "sum Available", "qty"])
	output_vendor_col = pick_column(output_df, ["Supplier", "Vendor", "Vendor name"])

	db_gtin_col = pick_column(db_df, ["Gtin", "GTIN", "EAN", "barcode"])
	db_hifi_col = pick_column(db_df, ["Hifi code", "Product code", "Unnamed: 1"])
	db_brand_col = pick_column(db_df, ["BrandName", "Brand", "Name"])
	db_title_col = pick_column(db_df, ["roductTitle", "ProductTitle", "Title"])

	output_df["_ean13"] = output_df[output_ean_col].apply(normalize_gtin)
	output_df["_price_num"] = pd.to_numeric(output_df[output_price_col], errors="coerce")
	output_df["_qty_num"] = pd.to_numeric(output_df[output_qty_col], errors="coerce")
	output_df["_stock_value"] = output_df["_price_num"] * output_df["_qty_num"]
	# Client rule: skip low total-value stock lines (< 100 EUR).
	output_df = output_df[output_df["_stock_value"] >= 100].copy()
	output_df = apply_supplier_filters(output_df, output_vendor_col, output_name_col)
	output_df = dedupe_by_ean_lowest_price(
		output_df,
		ean_col=output_ean_col,
		price_col=output_price_col,
		qty_col=output_qty_col,
	)

	db_df["_ean13"] = db_df[db_gtin_col].apply(normalize_gtin)

	db_lookup = db_df[
		["_ean13", db_hifi_col, db_brand_col, db_title_col]
	].drop_duplicates(subset=["_ean13"], keep="first")

	merged = output_df.merge(db_lookup, on="_ean13", how="left", indicator=True)
	last_prefix, last_number, width = get_last_product_code(db_df, db_hifi_col)

	unmatched_mask = merged["_merge"] == "left_only"
	unmatched_count = int(unmatched_mask.sum())
	new_codes = [
		f"{last_prefix}{str(last_number + i + 1).zfill(width)}"
		for i in range(unmatched_count)
	]
	if unmatched_count:
		merged.loc[unmatched_mask, db_hifi_col] = new_codes

	# Preserve the original output title column before assigning final Name.
	output_title = merged[output_name_col].copy()
	output_brand = output_title.apply(derive_brand)

	merged["barcode"] = merged["_ean13"]
	merged["Hifi code"] = merged[db_hifi_col].fillna("")
	merged["Name"] = merged[db_brand_col].where(
		merged[db_brand_col].notna() & (merged[db_brand_col].astype(str).str.strip() != ""),
		output_brand,
	)
	merged["ProductTitle"] = merged[db_title_col].where(
		merged[db_title_col].notna() & (merged[db_title_col].astype(str).str.strip() != ""),
		output_title,
	)
	merged["selling Price"] = pd.to_numeric(merged[output_price_col], errors="coerce").round(2)
	merged["sum Available"] = merged[output_qty_col]
	merged["Hifi price"] = merged.apply(
		lambda row: calculate_hifi_price_with_overrides(
			row["selling Price"],
			row["Name"],
			row["ProductTitle"],
		),
		axis=1,
	)
	merged["Vendor name"] = merged[output_vendor_col].fillna("")
	merged["Min quantity"] = merged.apply(
		lambda row: calculate_min_quantity(row["sum Available"], row["Hifi price"]),
		axis=1,
	)
	merged["price kole"] = (
		pd.to_numeric(merged["Min quantity"], errors="coerce")
		* pd.to_numeric(merged["Hifi price"], errors="coerce")
	).round(2)

	final_df = merged[
		[
			"barcode",
			"Hifi code",
			"Name",
			"ProductTitle",
			"selling Price",
			"sum Available",
			"Hifi price",
			"Vendor name",
			"Min quantity",
			"price kole",
		]
	].copy()

	unmatched = merged[merged["_merge"] == "left_only"].copy()
	new_products_df = unmatched[["barcode", "Hifi code", "Name", "ProductTitle"]].copy()
	new_products_df = new_products_df.rename(columns={"ProductTitle": "Product title"})

	with pd.ExcelWriter(FINAL_OUTPUT_FILE, engine="openpyxl") as writer:
		final_df.to_excel(writer, sheet_name="finaloutput", index=False)
		new_products_df.to_excel(writer, sheet_name="new products", index=False)

	bol_df = build_bol_dataframe(final_df)
	with pd.ExcelWriter(BOL_OUTPUT_FILE, engine="openpyxl") as writer:
		bol_df.to_excel(writer, sheet_name="Aanbod", index=False)

	digitec_df = build_digitec_dataframe(final_df)
	with pd.ExcelWriter(DIGITEC_OUTPUT_FILE, engine="openpyxl") as writer:
		digitec_df.to_excel(writer, sheet_name="digitec", index=False)

	apply_text_format_to_barcode(FINAL_OUTPUT_FILE)
	apply_text_format_to_barcode(BOL_OUTPUT_FILE)
	apply_text_format_to_barcode(DIGITEC_OUTPUT_FILE)
	print(f"Created: {FINAL_OUTPUT_FILE}")
	print(f"Created: {BOL_OUTPUT_FILE}")
	print(f"Created: {DIGITEC_OUTPUT_FILE}")


if __name__ == "__main__":
	main()
